#!/usr/bin/env python3
"""
Swim meet PDF (program / heats) -> Excel template generator

Updated behaviours
- Main 'Heats' sheet excludes Alternates entirely.
- All alternates are collected into a separate sheet: 'Alternates'.
- Gender: W for Girls/Women, M for Boys/Men, X for Mixed.
- Event shorthand: 50FS, 200BK, 100BR, 200FLY, 200IM, etc.
- Age group extracted into its own column between Event and Heat.
- Heat label uses the "1a Super Final" / "1b 15 Year Olds" part (prefix removed).
- Keeps your existing fast path / file identification behaviour from your script.
"""

from __future__ import annotations

import sys
import re
import datetime as dt
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple, Union, BinaryIO
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

AGE_PAT = re.compile(r"^(\d{1,2})$")

# Optional safety cap on how many heats/finals we keep per event.
# Set to None for unlimited.
# Example: MAX_HEATS_PER_EVENT = 3
MAX_HEATS_PER_EVENT: Optional[int] = None

@dataclass
class Heat:
    raw_label: str
    label: str
    lanes: Dict[int, str] = field(default_factory=dict)

@dataclass
class Event:
    number: int
    gender: str           # W/M/X
    event_code: str       # e.g. 50FS
    age_group: str        # e.g. 15 & Over
    heats: List[Heat] = field(default_factory=list)

@dataclass
class AlternateEntry:
    event_no: int
    gender: str
    event_code: str
    age_group: str
    heat_label: str       # which heat they belong to
    alt_group: str        # heading line after "Alternates ..."
    rank: int
    name: str
    team: str = ""
    prelim: str = ""

def normalise_name(name: str) -> str:
    """Normalise swimmer names.

    - Removes (V)
    - Removes trailing multi-class tokens like SM9 / SM10 / S14 / SB9 etc.
    - Normalises commas/whitespace
    - Uppercases
    """
    name = name.strip()

    # Remove common “visitor” marker.
    name = re.sub(r"\(\s*V\s*\)", "", name, flags=re.IGNORECASE)

    # Remove trailing multi-class codes, typically appended at end of the name.
    # Examples: SM9, SM10, SM19, S14, SB9
    name = re.sub(r"\s+S[A-Z]{0,2}\d{1,2}\s*$", "", name, flags=re.IGNORECASE)

    # Normalise punctuation/spacing
    name = name.replace(" ,", ",").replace(",", ", ")
    name = re.sub(r",\s+", ", ", name)
    name = re.sub(r"\s+", " ", name)

    return name.strip().upper()

def stroke_to_code(stroke: str) -> str:
    s = stroke.strip().lower()
    # common variants
    if "free" in s:
        return "FS"
    if "back" in s:
        return "BK"
    if "breast" in s:
        return "BR"
    if "butter" in s or "fly" in s:
        return "FLY"
    if "medley" in s or "im" == s:
        return "IM"
    # fallback: squeeze
    return re.sub(r"\s+", "", stroke).upper()

def parse_event_header(line: str) -> Optional[Tuple[int, str, str, str]]:
    """
    Example:
      Event 1  Girls 15 & Over 50 LC Meter Freestyle
    Returns:
      (1, 'W', '50FS', '15 & Over')
    """
    line = re.sub(r"\s+", " ", line.strip())
    m = re.match(r"^Event\s+(\d+)\s+(Girls|Women|Boys|Men|Mixed)\s+(.+)$", line, re.IGNORECASE)
    if not m:
        return None

    number = int(m.group(1))
    gender_word = m.group(2).lower()
    gender = {"girls":"W","women":"W","boys":"M","men":"M","mixed":"X"}.get(gender_word, "")

    rest = m.group(3).strip()

    # Find distance + stroke at end: "<dist> LC Meter <stroke>"
    m2 = re.search(r"(\d+)\s+LC\s+Meter\s+(.+)$", rest, re.IGNORECASE)
    if not m2:
        # fallback: try "Meter" without LC
        m2 = re.search(r"(\d+)\s+Meter\s+(.+)$", rest, re.IGNORECASE)
    if not m2:
        return number, gender, rest.upper(), ""  # worst-case fallback

    dist = m2.group(1)
    stroke_raw = re.sub(r"\s+", " ", m2.group(2).strip())

    # Multi-class events sometimes appear as e.g. "IM Multi-Class".
    # We want: "200IM MC" (not "200IMMULTI-CLASS").
    is_multiclass = bool(re.search(r"\bmulti\s*-?\s*class\b", stroke_raw, flags=re.IGNORECASE))

    # Remove the multi-class marker from the stroke descriptor before coding.
    stroke = re.sub(r"\bmulti\s*-?\s*class\b", "", stroke_raw, flags=re.IGNORECASE).strip()
    stroke = re.sub(r"\s+", " ", stroke)

    age_group = rest[:m2.start()].strip()  # everything before distance
    event_code = f"{dist}{stroke_to_code(stroke)}" + (" MC" if is_multiclass else "")
    return number, gender, event_code, age_group

def clean_heat_label(label: str) -> str:
    """Remove age-group-ish fragments from a heat label.

    We keep the heat identifier (e.g. "5a") and any non-age descriptor
    (e.g. "Super Final"), but strip age phrases that some programs append.

    Examples:
      "1b 15 Year Olds" -> "1b"
      "2 10 & Over" -> "2"
      "2a 12-13 Years & Old" -> "2a"
      "5a 12" -> "5a"
      "1a Super Final" -> "1a Super Final" (unchanged)
    """
    s = re.sub(r"\s+", " ", label.strip())

    # Normalise unicode dashes into '-' for easier matching.
    s = s.replace("–", "-").replace("—", "-")

    # IMPORTANT: range patterns must run BEFORE single-age patterns.
    # Otherwise e.g. "12-13 Years Olds" might match the "13 Years Olds" tail first,
    # leaving behind a stray "12".
    patterns = [
        # 12-13 Years & Old / 12-16 Years Olds / 12 - 13 Years Old
        r"\b\d{1,2}\s*-\s*\d{1,2}\s*Years?\s*(?:&|and)?\s*Olds?\b",

        # Some PDFs render as: "12-16 Years" (no Old/Over word)
        r"\b\d{1,2}\s*-\s*\d{1,2}\s*Years?\b",

        # 17 Years & Over / 17 Years and Over
        r"\b\d{1,2}\s*Years?\s*(?:&|and)\s*(?:Over|Under)\b",
        r"\b\d{1,2}\s*(?:&|and)\s*(?:Over|Under)\b",

        # 15 Year Olds / 15 Years Old
        r"\b\d{1,2}\s*Years?\s*Olds?\b",

        # Some PDFs render as: "Years & Over" without the preceding number token
        r"\bYears?\s*(?:&|and)\s*(?:Over|Under)\b",
    ]

    for pat in patterns:
        s = re.sub(pat, "", s, flags=re.IGNORECASE)

    s = re.sub(r"\s+", " ", s).strip()

    # Clean up trailing separators early (so bare-age removal can match).
    s = s.rstrip("- ").strip()

    # If label looks like "3a 12" (heat + bare age), drop the trailing age.
    s = re.sub(r"^(\S+)\s+\d{1,2}$", r"\1", s)

    return s.strip()


def parse_heat_label(line: str) -> Optional[str]:
    """
    "Final  1a  Super Final" -> "1a Super Final"
    "Final  1b  15 Year Olds" -> "1b"
    "Heat 2" -> "2"
    """
    line = re.sub(r"\s+", " ", line.strip())
    m = re.match(r"^(Final|Heat)\s+(.+)$", line, re.IGNORECASE)
    if not m:
        return None
    return clean_heat_label(m.group(2).strip())

def parse_lane_line(line: str) -> Optional[Tuple[int, str]]:
    line = line.strip()
    m = re.match(r"^([0-9])\s+(.*)$", line)
    if not m:
        return None
    lane = int(m.group(1))
    rest = m.group(2).strip()
    tokens = rest.split()

    # Multi-class codes like SM9/SM10/S14 etc.
    mc_pat = re.compile(r"^S[A-Z]{0,2}\d{1,2}$", re.IGNORECASE)
    # Sex+age tokens sometimes included in exports (e.g. W17 / M15)
    sex_age_pat = re.compile(r"^[MWX]\d{1,2}$", re.IGNORECASE)

    name_tokens: List[str] = []
    for tok in tokens:
        if AGE_PAT.match(tok):
            break
        if sex_age_pat.match(tok):
            break
        if mc_pat.match(tok):
            break
        if re.match(r"^(NT|\d{1,2}:\d{2}\.\d{2})$", tok, re.IGNORECASE):
            break
        name_tokens.append(tok)

    if not name_tokens:
        return None

    name = " ".join(name_tokens).strip()
    if not name:
        return None
    return lane, normalise_name(name)

def parse_alternate_line(line: str) -> Optional[Tuple[int, str, str, str]]:
    """
    Alternate lines look like:
      1 Shumack, Heidi 16 Sopac 26.25
      2 Hamilton (V), Nafanua 15 Samoa 27.74
    Returns: (rank, NAME, TEAM, PRELIMS)
    """
    line = re.sub(r"\s+", " ", line.strip())
    m = re.match(r"^(\d+)\s+(.*)$", line)
    if not m:
        return None
    rank = int(m.group(1))
    rest = m.group(2).strip()
    tokens = rest.split()

    # name tokens up to first standalone age (number)
    name_tokens: List[str] = []
    idx_age = None
    mc_pat = re.compile(r"^S[A-Z]{0,2}\d{1,2}$", re.IGNORECASE)
    sex_age_pat = re.compile(r"^[MWX]\d{1,2}$", re.IGNORECASE)

    for i, tok in enumerate(tokens):
        if AGE_PAT.match(tok):
            idx_age = i
            break
        # sometimes there are sex+age tokens like W17/M15 in some programs
        if sex_age_pat.match(tok):
            idx_age = i
            break
        # multi-class code embedded in alternates lists
        if mc_pat.match(tok):
            idx_age = i
            break
        name_tokens.append(tok)

    if not name_tokens:
        return None

    name = normalise_name(" ".join(name_tokens))

    team = ""
    prelim = ""

    if idx_age is not None:
        # remaining tokens after age token
        rem = tokens[idx_age+1:]
        # prelim time is usually last time-like token
        for j, tok in enumerate(rem):
            if re.match(r"^\d{1,2}:\d{2}\.\d{2}$|^\d{1,2}\.\d{2}$|^NT$", tok, re.IGNORECASE):
                # team is tokens before this, prelim is this token
                team = " ".join(rem[:j]).strip()
                prelim = tok
                break
        if not prelim and rem:
            # if nothing matched, assume last token is prelim
            prelim = rem[-1]
            team = " ".join(rem[:-1]).strip()

    return rank, name, team.upper(), prelim

def _discover_pdfs_in_cwd() -> List[Path]:
    """Return PDFs in the current working directory.

    Raises ValueError if none are found.
    """
    pdfs = sorted(Path.cwd().glob("*.pdf"), key=lambda p: p.name.lower())
    if not pdfs:
        raise ValueError("No PDFs found in the current directory")
    return pdfs


def _prompt_user_to_select_pdf(pdfs: List[Path]) -> Path:
    """Ask the user which PDF to parse.

    If only one PDF exists, selects it automatically.
    Raises ValueError for invalid selections.
    """
    if not pdfs:
        raise ValueError("No PDFs available to select")

    if len(pdfs) == 1:
        return pdfs[0]

    print("Found PDFs:")
    for i, p in enumerate(pdfs, start=1):
        print(f"  {i}) {p.name}")

    choice = input(f"Select PDF [1-{len(pdfs)}]: ").strip()
    if not choice.isdigit():
        raise ValueError("Selection must be a number")

    idx = int(choice)
    if idx < 1 or idx > len(pdfs):
        raise ValueError("Selection out of range")

    return pdfs[idx - 1]


def infer_day_title(first_page_lines: List[str]) -> str:
    meet_line = first_page_lines[0] if first_page_lines else ""
    program_line = first_page_lines[1] if len(first_page_lines) > 1 else ""

    start_date = None
    m = re.search(r"-\s*(\d{1,2}/\d{1,2}/\d{4})\s+to\s+(\d{1,2}/\d{1,2}/\d{4})", meet_line)
    if m:
        start_date = dt.datetime.strptime(m.group(1), "%d/%m/%Y").date()

    night_no = 1
    m2 = re.search(r"Night\s+(One|Two|Three|Four|Five|Six|Seven|Eight|Nine|Ten|\d+)", program_line, re.IGNORECASE)
    if m2:
        token = m2.group(1).lower()
        words = {"one":1,"two":2,"three":3,"four":4,"five":5,"six":6,"seven":7,"eight":8,"nine":9,"ten":10}
        night_no = words.get(token, int(token) if token.isdigit() else 1)

    date = start_date + dt.timedelta(days=night_no-1) if start_date else None
    date_str = date.strftime("%d/%m/%Y") if date else ""
    return f"Day {night_no} Heats - {date_str}"

def parse_pdf(pdf_source: Union[str, Path, BinaryIO]) -> Tuple[str, List[Event], List[AlternateEntry]]:
    """Parse a meet program PDF into structured events + alternates.

    `pdf_source` can be:
      - a filesystem path (str/Path) (CLI usage)
      - a file-like object / BytesIO (Streamlit uploads)

    pdfplumber can open both.
    """
    with pdfplumber.open(pdf_source) as pdf:
        pages = [(p.extract_text() or "") for p in pdf.pages]

    title = infer_day_title(pages[0].splitlines())

    events: List[Event] = []
    alternates: List[AlternateEntry] = []

    current_event: Optional[Event] = None
    current_heat: Optional[Heat] = None
    in_alternates = False
    current_alt_group = ""

    for page_text in pages:
        for raw in page_text.splitlines():
            line = raw.strip()
            if not line:
                continue

            # new event
            ev = parse_event_header(line)
            if ev:
                if current_event:
                    events.append(current_event)
                number, gender, event_code, age_group = ev
                current_event = Event(number=number, gender=gender, event_code=event_code, age_group=age_group)
                current_heat = None
                in_alternates = False
                current_alt_group = ""
                continue

            if not current_event:
                continue

            low = line.lower()

            # skip boilerplate
            if low.startswith("lane ") or low.startswith("name ") or low.startswith("age ") or low.startswith("team "):
                continue
            if low.startswith("finals program"):
                continue
            if re.match(r"^\d{4}-\d{2}\b", line):
                continue

            # alternates heading
            if low.startswith("alternates"):
                in_alternates = True
                current_alt_group = re.sub(r"\s+", " ", line.strip())
                # stop collecting lanes into heat while in alternates
                continue

            # new heat start ends alternates mode
            heat_label = parse_heat_label(line)
            if heat_label is not None:
                in_alternates = False
                current_alt_group = ""
                # Optional cap to avoid runaway parsing if a PDF is malformed.
                if MAX_HEATS_PER_EVENT is not None and len(current_event.heats) >= MAX_HEATS_PER_EVENT:
                    current_heat = None
                else:
                    current_heat = Heat(raw_label=line, label=heat_label)
                    current_event.heats.append(current_heat)
                continue

            # while in alternates: collect alternate lines, but do NOT treat as lanes
            if in_alternates:
                parsed_alt = parse_alternate_line(line)
                if parsed_alt and current_event and current_event.heats:
                    rank, name, team, prelim = parsed_alt
                    alternates.append(
                        AlternateEntry(
                            event_no=current_event.number,
                            gender=current_event.gender,
                            event_code=current_event.event_code,
                            age_group=current_event.age_group,
                            heat_label=current_event.heats[-1].label,
                            alt_group=current_alt_group,
                            rank=rank,
                            name=name,
                            team=team,
                            prelim=prelim,
                        )
                    )
                continue

            # normal lane row
            parsed = parse_lane_line(line)
            if parsed and current_heat:
                lane, name = parsed
                current_heat.lanes[lane] = name

    if current_event:
        events.append(current_event)

    return title, events, alternates

def build_workbook(events: List[Event], alternates: List[AlternateEntry], title: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Heats"

    headers = ["#", "Gender", "Event", "Age Group", "Heat", "Cal"] + [f"Lane {i}" for i in range(10)] + [f"Analyst {i}" for i in range(1, 5)]
    ncols = len(headers)

    # Row 1 title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Row 2 headers
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    pastel = ["FFF2CC", "DDEBF7", "E2F0D9"]

    row = 3
    for idx, ev in enumerate(events):
        fill = PatternFill("solid", fgColor=pastel[idx % 3])
        start_row = row

        for heat in ev.heats:
            ws.cell(row=row, column=1, value=ev.number).fill = fill
            ws.cell(row=row, column=2, value=ev.gender).fill = fill
            ws.cell(row=row, column=3, value=ev.event_code).fill = fill
            ws.cell(row=row, column=4, value=ev.age_group).fill = fill
            ws.cell(row=row, column=5, value=heat.label).fill = fill
            ws.cell(row=row, column=6, value="").fill = fill  # Cal blank

            # lanes
            for lane in range(10):
                col = 7 + lane
                ws.cell(row=row, column=col, value=heat.lanes.get(lane, "")).fill = fill

            # analysts blank
            for a in range(4):
                ws.cell(row=row, column=17 + a, value="").fill = fill

            row += 1

        end_row = row - 1
        if end_row >= start_row:
            # merge event-level columns across heats
            for col in [1,2,3,4]:
                ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                ws.cell(row=start_row, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # borders + align
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ncols):
        for cell in r:
            cell.border = border
            if cell.row >= 3 and cell.column >= 7:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # column widths
    widths = {
        1: 5, 2: 8, 3: 10, 4: 14, 5: 18, 6: 6
    }
    for i in range(10):
        widths[7+i] = 22
    for i in range(4):
        widths[17+i] = 10

    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Alternates sheet
    ws2 = wb.create_sheet("Alternates")
    alt_headers = ["#", "Gender", "Event", "Age Group", "Heat", "Alt Group", "Alt Rank", "Name", "Team", "Prelims"]
    alt_ncols = len(alt_headers)

    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=alt_ncols)
    c2 = ws2.cell(row=1, column=1, value=title + " (Alternates)")
    c2.font = Font(bold=True, size=14)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 22

    for col, h in enumerate(alt_headers, start=1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill
    ws2.freeze_panes = "A3"

    r = 3
    for a in alternates:
        ws2.cell(r,1,a.event_no)
        ws2.cell(r,2,a.gender)
        ws2.cell(r,3,a.event_code)
        ws2.cell(r,4,a.age_group)
        ws2.cell(r,5,a.heat_label)
        ws2.cell(r,6,a.alt_group)
        ws2.cell(r,7,a.rank)
        ws2.cell(r,8,a.name)
        ws2.cell(r,9,a.team)
        ws2.cell(r,10,a.prelim)
        r += 1

    for row_cells in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=alt_ncols):
        for cell in row_cells:
            cell.border = border
            if cell.column in (8,9,6):
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 18
    ws2.column_dimensions["F"].width = 22
    ws2.column_dimensions["G"].width = 10
    ws2.column_dimensions["H"].width = 26
    ws2.column_dimensions["I"].width = 18
    ws2.column_dimensions["J"].width = 10

    return wb

def main() -> int:
    # Legacy mode: explicit input/output args
    if len(sys.argv) == 3:
        pdf_path = sys.argv[1]
        out_xlsx = sys.argv[2]
    # Interactive mode: discover and prompt
    elif len(sys.argv) == 1:
        try:
            pdf = _prompt_user_to_select_pdf(_discover_pdfs_in_cwd())
        except ValueError as e:
            print(str(e))
            print("Tip: run this script from a folder containing PDFs, or provide paths explicitly:")
            print("  python pdf_to_heats_xlsx.py /path/to/program.pdf /path/to/output.xlsx")
            return 2

        pdf_path = str(pdf)
        out_xlsx = str(pdf.with_suffix(".xlsx"))
        print(f"Selected: {pdf_path}")
        print(f"Output:   {out_xlsx}")
    else:
        print("Usage:")
        print("  python pdf_to_heats_xlsx.py")
        print("  python pdf_to_heats_xlsx.py /path/to/program.pdf /path/to/output.xlsx")
        return 2

    title, events, alternates = parse_pdf(pdf_path)
    wb = build_workbook(events, alternates, title)
    wb.save(out_xlsx)
    print(f"Saved: {out_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
